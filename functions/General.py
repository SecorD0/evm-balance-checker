import logging
import os
import random
from typing import List, Dict, Any

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from py_eth_async.client import Client
from py_eth_async.data.models import Network

from data import config
from data.models import Settings
from functions.get_network_name import get_network_name
from functions.get_random_proxy import get_random_proxy
from functions.get_token_price import get_token_price
from utils.db_api.database import get_address, db, get_addresses


class General:
    @staticmethod
    async def import_addresses(settings: Settings, addresses: List[Dict[str, Any]]) -> None:
        try:
            if addresses:
                print('Importing addresses...')
                imported = []
                total = len(addresses)
                for network in settings.networks:
                    try:
                        columns = ''
                        evm_network: Network = random.choice(network.rpcs)
                        chain_id = evm_network.chain_id
                        client = Client(private_key='', network=evm_network, proxy=await get_random_proxy())
                        if network.tokens:
                            for token in network.tokens:
                                if token:
                                    contract = await client.contracts.default_token(contract_address=token)
                                    token_symbol = await contract.functions.symbol().call()
                                    token_symbol = token_symbol.lower()
                                    columns += f"'{token_symbol}' REAL, "

                                else:
                                    columns += f"'eth' REAL, "

                        else:
                            columns += "'eth' REAL, "

                        if network.nfts:
                            for nft in network.nfts.values():
                                contract = await client.contracts.default_token(contract_address=nft.contract_address)
                                token_symbol = await contract.functions.symbol().call()
                                token_symbol = token_symbol.lower()
                                column_name = f'{token_symbol.replace(" ", "_")}'
                                if nft.token_id is not None:
                                    column_name += f'_{nft.token_id}'

                                columns += f"'{column_name}' INTEGER, "

                        columns = columns[:-2]
                        db.execute(f"CREATE TABLE IF NOT EXISTS '{chain_id}' (address TEXT PRIMARY KEY, {columns})")
                        for address in addresses:
                            try:
                                address = address.get('address')
                                if address and not get_address(chain_id=chain_id, address=address):
                                    db.execute(f"INSERT INTO '{chain_id}' (address) VALUES (?)", (address,))
                                    if address not in imported:
                                        imported.append(address)

                            except BaseException as e:
                                logging.exception('General.import_addresses (address)')
                                print(f'{config.RED}Failed to import an address: {e}{config.RESET_ALL}')

                    except BaseException as e:
                        logging.exception('General.import_addresses (network)')
                        print(f'{config.RED}Failed to use network to import an address: {e}{config.RESET_ALL}')

                print(f'Done! {config.LIGHTGREEN_EX}{len(imported)}/{total}{config.RESET_ALL} addresses were imported.')

            else:
                print(f"{config.RED}You didn't provide any addresses!{config.RESET_ALL}")

        except BaseException as e:
            logging.exception('General.import_addresses')
            print(f"{config.RED}Failed to import addresses: {e}{config.RESET_ALL}")

    @staticmethod
    async def export_addresses(settings: Settings) -> None:
        try:
            spreadsheet = Workbook()
            for i, network in enumerate(settings.networks):
                evm_network: Network = random.choice(network.rpcs)
                chain_id = evm_network.chain_id
                addresses = get_addresses(chain_id=chain_id)
                if addresses:
                    if i:
                        sheet: Worksheet = spreadsheet.create_sheet(get_network_name(network=network))

                    else:
                        sheet: Worksheet = spreadsheet['Sheet']
                        sheet.title = get_network_name(network=network)

                    token_symbols = []
                    token_prices = {}
                    if settings.general.parse_token_price:
                        if network.tokens:
                            client = Client(private_key='', network=evm_network, proxy=await get_random_proxy())
                            for token in network.tokens:
                                if token:
                                    contract = await client.contracts.default_token(contract_address=token)
                                    token_symbol = await contract.functions.symbol().call()
                                    token_symbols.append(token_symbol.lower())

                                else:
                                    token_symbols.append('eth')

                        else:
                            token_symbols.append('eth')

                        token_prices = await get_token_price(token_symbols=token_symbols)

                    headers = ['n']
                    db_headers = [
                        header[1] for header in db.execute(f"PRAGMA table_info({chain_id})", return_class=False)
                    ]
                    for header in db_headers:
                        headers.append(header)
                        if header in token_symbols:
                            headers.append(f'{header}_usd')

                    if settings.general.parse_token_price:
                        headers.append('total_balance_usd')

                    db_rows: List[Dict[str, Any]] = []
                    for address in addresses:
                        db_rows.append(dict(zip(db_headers, address)))

                    rows: List[Dict[str, Any]] = []
                    for j, db_row in enumerate(db_rows):
                        total_balance_usd = 0
                        row = {}
                        for header in headers:
                            if header == 'n':
                                row[header] = j + 1

                            elif header in db_row:
                                row[header] = db_row[header]

                            elif settings.general.parse_token_price and '_usd' in header:
                                token = header.replace('_usd', '')
                                if token in token_prices:
                                    row[header] = db_row[token] * token_prices[token]
                                    total_balance_usd += row[header]

                                else:
                                    row[header] = None

                        if settings.general.parse_token_price:
                            row['total_balance_usd'] = total_balance_usd

                        rows.append(row)

                    for column, header in enumerate(headers):
                        sheet.cell(row=1, column=column + 1).value = header

                    for n, row in enumerate(rows):
                        for column, value in enumerate(row.values()):
                            cell = sheet.cell(row=n + 2, column=column + 1)
                            cell.value = value

            try:
                file = open(file=config.ADDRESSES_FILE, mode='r+')
                spreadsheet.save(config.ADDRESSES_FILE)
                db.db.close()
                os.remove(path=config.ADDRESSES_DB)
                print(
                    f'\nThe results were saved to the '
                    f'{config.LIGHTGREEN_EX}{config.ADDRESSES_FILE}{config.RESET_ALL} file.'
                )

            except IOError:
                print(f"\n{config.RED}You didn't close the {config.ADDRESSES_FILE} file!{config.RESET_ALL}")

        except BaseException as e:
            logging.exception('General.export_addresses')
            print(f"{config.RED}Failed to export addresses: {e}{config.RESET_ALL}")
